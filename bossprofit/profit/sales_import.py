import hashlib
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
from statistics import median

from openpyxl import load_workbook


EXPECTED_HEADERS = (
    "대분류",
    "상품코드",
    "상품명",
    "일자",
    "수량",
    "총매출액",
    "총할인액",
    "실매출액",
)


@dataclass(frozen=True)
class PosSaleRow:
    category: str
    product_code: str
    product_name: str
    sale_date: date
    quantity: int
    gross_revenue: int
    discount_amount: int
    net_revenue: int


def sha256_file(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_pos_workbook(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    for row in rows:
        normalized = tuple("" if value is None else str(value).strip() for value in row[:8])
        if normalized == EXPECTED_HEADERS:
            break
    else:
        raise ValueError(f"POS 헤더를 찾을 수 없습니다: {Path(path).name}")

    current_category = None
    current_product_code = None
    current_product_name = None
    parsed = []
    for row in rows:
        if not any(value is not None for value in row):
            continue
        if row[0] is not None:
            category = str(row[0]).strip()
            if category == "합계":
                continue
            current_category = category
        if row[1] is not None and str(row[1]).strip():
            current_product_code = str(row[1]).strip()
        if row[2] is not None and str(row[2]).strip():
            current_product_name = str(row[2]).strip()
        if (
            not current_category
            or not current_product_code
            or not current_product_name
            or row[3] is None
        ):
            continue
        sale_date = row[3].date() if hasattr(row[3], "date") else row[3]
        parsed.append(
            PosSaleRow(
                category=current_category,
                product_code=current_product_code,
                product_name=current_product_name,
                sale_date=sale_date,
                quantity=max(0, int(Decimal(str(row[4] or 0)))),
                gross_revenue=max(0, int(Decimal(str(row[5] or 0)))),
                discount_amount=max(0, int(Decimal(str(row[6] or 0)))),
                net_revenue=max(0, int(Decimal(str(row[7] or 0)))),
            )
        )
    workbook.close()
    return parsed


def summarize_product_prices(rows):
    prices = defaultdict(list)
    for row in rows:
        if row.quantity > 0 and row.gross_revenue > 0:
            prices[row.product_code].append(row.gross_revenue / row.quantity)
    return {
        product_code: int(round(median(values)))
        for product_code, values in prices.items()
        if values
    }
